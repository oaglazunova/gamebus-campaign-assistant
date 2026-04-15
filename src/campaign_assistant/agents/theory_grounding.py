from __future__ import annotations

from pathlib import Path
from typing import Any

from campaign_assistant.agents.base import BaseAgent
from campaign_assistant.orchestration.models import AgentContext, AgentResponse


TTM_STAGE_EXPLANATIONS = {
    "precontemplation": "Focus is typically on awareness, reflection, and risk/benefit understanding rather than action-heavy demands.",
    "contemplation": "Intervention elements often support reflection, pros/cons weighing, and motivation to change.",
    "preparation": "This stage usually benefits from planning, goal-setting, and concrete preparation for behavior change.",
    "action": "The focus shifts toward doing the behavior, self-monitoring, feedback, and sustaining execution.",
    "maintenance": "The intervention should support continuation, relapse prevention, reinforcement, and habit stabilization.",
}


def _resolve_workspace_relative_path(
    workspace_root: Path | None,
    maybe_relative_path: str | None,
) -> Path | None:
    if not maybe_relative_path:
        return None

    raw = Path(maybe_relative_path)
    if raw.is_absolute():
        return raw

    if workspace_root is None:
        return None

    return workspace_root / raw


def _safe_list_task_roles(task_roles: list[dict[str, str]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in task_roles:
        role = (row.get("role") or "").strip().lower()
        if not role:
            continue
        counts[role] = counts.get(role, 0) + 1
    return counts


def _load_mapping_file_summary(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {"exists": False, "reason": "No mapping file configured"}

    if not path.exists():
        return {"exists": False, "reason": f"Missing file: {path}"}

    summary: dict[str, Any] = {"exists": True, "path": str(path)}

    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        summary["sheet_names"] = wb.sheetnames
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
            summary["first_sheet_title"] = ws.title
            summary["first_sheet_max_row"] = ws.max_row
            summary["first_sheet_max_column"] = ws.max_column
        wb.close()
    except Exception as exc:
        summary["metadata_warning"] = f"Could not inspect workbook details: {exc}"

    return summary


class TheoryGroundingAgent(BaseAgent):
    name = "theory_grounding_agent"

    def run(self, context: AgentContext) -> AgentResponse:
        profile = context.analysis_profile or {}
        checking_scope = profile.get("checking_scope", {})
        intervention_model = profile.get("intervention_model", {})
        ttm_cfg = profile.get("ttm", {})
        theory_cfg = profile.get("theory_grounding", {})

        if not checking_scope.get("theory_checks", False):
            payload = {
                "theory_checks_enabled": False,
                "reason": "Theory checks are disabled in the analysis profile.",
            }
            context.shared["theory_grounding"] = payload
            return AgentResponse(
                agent_name=self.name,
                success=True,
                summary="Theory grounding skipped because theory checks are disabled in the workspace profile.",
                payload=payload,
            )

        uses_ttm = bool(intervention_model.get("uses_ttm", False))
        uses_bct = bool(intervention_model.get("uses_bct_mapping", False))
        uses_comb = bool(intervention_model.get("uses_comb_mapping", False))

        ttm_file = _resolve_workspace_relative_path(
            context.workspace_root,
            ttm_cfg.get("structure_file"),
        )
        mapping_file = _resolve_workspace_relative_path(
            context.workspace_root,
            theory_cfg.get("mapped_interventions_file"),
        )

        ttm_file_exists = bool(ttm_file and ttm_file.exists())
        mapping_summary = _load_mapping_file_summary(mapping_file)
        role_counts = _safe_list_task_roles(context.task_roles)

        structural_result = context.shared.get("result", {})
        structural_summary = structural_result.get("summary", {})
        failed_checks = structural_summary.get("failed_checks", [])

        point_gatekeeping = structural_result.get("point_gatekeeping", {})
        pg_summary = point_gatekeeping.get("summary", {})
        pg_findings = pg_summary.get("challenge_findings", 0)
        pg_gatekeeper_warnings = pg_summary.get("gatekeeper_warnings", 0)
        pg_maintenance_warnings = pg_summary.get("maintenance_warnings", 0)

        warnings: list[str] = []
        notes: list[str] = []

        if uses_ttm and not ttm_file_exists:
            warnings.append(
                "The analysis profile says this campaign uses TTM, but the configured TTM structure file is missing."
            )

        if uses_bct or uses_comb:
            if not mapping_summary.get("exists", False):
                warnings.append(
                    "BCT/COM-B explanatory grounding is enabled in the profile, but the mapped interventions resource is missing."
                )

        if uses_ttm and "ttm" in failed_checks:
            notes.append(
                "The structural checker reported a TTM-related failure. "
                "This suggests the campaign path or level progression may not match the intended stage-of-change structure."
            )

        if intervention_model.get("uses_gatekeeping", False):
            if not role_counts:
                notes.append(
                    "No explicit task roles were provided. Gatekeeping and maintenance interpretations currently rely partly on inference."
                )

            if pg_gatekeeper_warnings:
                notes.append(
                    f"Point/gatekeeping analysis raised {pg_gatekeeper_warnings} gatekeeping warning(s), suggesting that progression logic may not fully reflect the intended intervention structure."
                )

            if pg_maintenance_warnings:
                notes.append(
                    f"Point/gatekeeping analysis raised {pg_maintenance_warnings} maintenance-task warning(s), which may matter for at-risk or relapse paths."
                )

        stage_notes: dict[str, str] = {}
        for stage in ttm_cfg.get("stages", []):
            key = str(stage).strip().lower()
            if key in TTM_STAGE_EXPLANATIONS:
                stage_notes[key] = TTM_STAGE_EXPLANATIONS[key]

        confidence = "high"
        if warnings:
            confidence = "low"
        elif not role_counts or pg_findings:
            confidence = "medium"

        summary_parts: list[str] = []
        if uses_ttm:
            if "ttm" in failed_checks:
                summary_parts.append(
                    "Theory grounding detected a mismatch between the intended TTM-based intervention structure and the structural checker outcome."
                )
            else:
                summary_parts.append(
                    "Theory grounding found no direct TTM structural conflict in the current checker results."
                )
        else:
            summary_parts.append(
                "Theory grounding is active, but this workspace profile does not mark TTM as operational for this campaign."
            )

        if pg_findings:
            summary_parts.append(
                f"Point/gatekeeping reasoning contributed {pg_findings} intervention-mechanics finding(s)."
            )

        if uses_bct or uses_comb:
            if mapping_summary.get("exists", False):
                summary_parts.append(
                    "BCT/COM-B grounding resources are available for explanatory support."
                )
            else:
                summary_parts.append(
                    "BCT/COM-B grounding was requested in the profile, but the mapping resource is not available."
                )

        payload = {
            "theory_checks_enabled": True,
            "uses_ttm": uses_ttm,
            "uses_bct_mapping": uses_bct,
            "uses_comb_mapping": uses_comb,
            "ttm_structure_file": str(ttm_file) if ttm_file else None,
            "ttm_structure_file_exists": ttm_file_exists,
            "mapping_summary": mapping_summary,
            "task_role_counts": role_counts,
            "point_gatekeeping_summary": pg_summary,
            "confidence": confidence,
            "notes": notes,
            "stage_notes": stage_notes,
            "failed_checks_seen": failed_checks,
        }

        context.shared["theory_grounding"] = payload

        return AgentResponse(
            agent_name=self.name,
            success=True,
            summary=" ".join(summary_parts),
            payload=payload,
            warnings=warnings,
        )