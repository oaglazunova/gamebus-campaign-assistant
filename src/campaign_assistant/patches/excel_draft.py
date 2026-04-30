from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class PatchedExcelDraftGenerator:
    """
    Generate a patched Excel draft from a patch manifest.

    Current scope:
    - create a copy of the analyzed campaign export
    - apply supported operations
    - add a notes sheet summarizing applied and unresolved operations
    - persist a JSON summary next to the draft

    Supported operations now:
    - set_target_points

    Unsupported operations are kept as unresolved/manual follow-up items.
    """

    NOTES_SHEET_NAME = "__assistant_patch_notes"

    def __init__(self, workspace_root: str | Path, request_id: str):
        self.workspace_root = Path(workspace_root)
        self.request_id = request_id
        self._patches_dir = self.workspace_root / "outputs" / "patches"
        self._patches_dir.mkdir(parents=True, exist_ok=True)

    @property
    def draft_path(self) -> Path:
        return self._patches_dir / f"{self.request_id}_patched_campaign.xlsx"

    @property
    def summary_path(self) -> Path:
        return self._patches_dir / f"{self.request_id}_patched_campaign_summary.json"

    def generate(
        self,
        *,
        snapshot_path: str | Path,
        manifest: dict[str, Any],
    ) -> dict[str, Any]:
        snapshot_path = Path(snapshot_path)

        if not snapshot_path.exists():
            raise FileNotFoundError(f"Snapshot file does not exist: {snapshot_path}")

        wb = load_workbook(snapshot_path)
        applied: list[dict[str, Any]] = []
        unresolved: list[dict[str, Any]] = []

        challenges_ws = wb["challenges"] if "challenges" in wb.sheetnames else None
        challenge_index = self._index_challenges(challenges_ws) if challenges_ws is not None else {}

        for op in manifest.get("operations", []):
            op_name = op.get("op")
            if op_name == "set_target_points":
                outcome = self._apply_set_target_points(
                    ws=challenges_ws,
                    challenge_index=challenge_index,
                    operation=op,
                )
                if outcome["status"] == "applied":
                    applied.append(outcome)
                else:
                    unresolved.append(outcome)
            else:
                unresolved.append(
                    {
                        "status": "unresolved",
                        "proposal_id": op.get("proposal_id"),
                        "op": op_name,
                        "challenge_name": op.get("challenge_name"),
                        "reason": "This operation is not yet executable in the Excel draft generator.",
                        "params": op.get("params", {}),
                    }
                )

        self._write_notes_sheet(
            wb=wb,
            manifest=manifest,
            applied=applied,
            unresolved=unresolved,
            snapshot_path=snapshot_path,
        )

        wb.save(self.draft_path)
        wb.close()

        summary = {
            "request_id": self.request_id,
            "draft_path": str(self.draft_path),
            "source_snapshot_path": str(snapshot_path),
            "manifest_operation_count": manifest.get("operation_count", 0),
            "applied_count": len(applied),
            "unresolved_count": len(unresolved),
            "applied": applied,
            "unresolved": unresolved,
        }

        with self.summary_path.open("w", encoding="utf-8") as fh:
            json.dump(summary, fh, indent=2, ensure_ascii=False)

        return summary

    def _index_challenges(self, ws) -> dict[str, int]:
        """
        Build a map: challenge_name -> row_index
        """
        if ws is None:
            return {}

        headers = self._headers(ws)
        if "name" not in headers:
            return {}

        name_col = headers["name"]
        challenge_index: dict[str, int] = {}

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=name_col).value
            if value is None:
                continue
            challenge_name = str(value).strip()
            if challenge_name and challenge_name not in challenge_index:
                challenge_index[challenge_name] = row_idx

        return challenge_index

    def _apply_set_target_points(
        self,
        *,
        ws,
        challenge_index: dict[str, int],
        operation: dict[str, Any],
    ) -> dict[str, Any]:
        if ws is None:
            return {
                "status": "unresolved",
                "proposal_id": operation.get("proposal_id"),
                "op": operation.get("op"),
                "challenge_name": operation.get("challenge_name"),
                "reason": "Workbook has no 'challenges' sheet.",
                "params": operation.get("params", {}),
            }

        headers = self._headers(ws)
        if "target" not in headers:
            return {
                "status": "unresolved",
                "proposal_id": operation.get("proposal_id"),
                "op": operation.get("op"),
                "challenge_name": operation.get("challenge_name"),
                "reason": "The 'challenges' sheet has no 'target' column.",
                "params": operation.get("params", {}),
            }

        challenge_name = operation.get("challenge_name")
        if not challenge_name:
            return {
                "status": "unresolved",
                "proposal_id": operation.get("proposal_id"),
                "op": operation.get("op"),
                "challenge_name": challenge_name,
                "reason": "Operation has no challenge_name.",
                "params": operation.get("params", {}),
            }

        row_idx = challenge_index.get(str(challenge_name).strip())
        if row_idx is None:
            return {
                "status": "unresolved",
                "proposal_id": operation.get("proposal_id"),
                "op": operation.get("op"),
                "challenge_name": challenge_name,
                "reason": "Challenge name could not be matched in the workbook.",
                "params": operation.get("params", {}),
            }

        target_col = headers["target"]
        previous_value = ws.cell(row=row_idx, column=target_col).value
        new_value = (operation.get("params") or {}).get("target_points")

        ws.cell(row=row_idx, column=target_col).value = new_value

        return {
            "status": "applied",
            "proposal_id": operation.get("proposal_id"),
            "op": operation.get("op"),
            "challenge_name": challenge_name,
            "previous_value": previous_value,
            "new_value": new_value,
        }

    def _write_notes_sheet(
        self,
        *,
        wb,
        manifest: dict[str, Any],
        applied: list[dict[str, Any]],
        unresolved: list[dict[str, Any]],
        snapshot_path: Path,
    ) -> None:
        if self.NOTES_SHEET_NAME in wb.sheetnames:
            del wb[self.NOTES_SHEET_NAME]

        ws = wb.create_sheet(self.NOTES_SHEET_NAME)

        row = 1
        ws.cell(row=row, column=1).value = "GameBus Campaign Assistant Patch Notes"
        row += 2

        ws.cell(row=row, column=1).value = "Source snapshot"
        ws.cell(row=row, column=2).value = str(snapshot_path)
        row += 1

        ws.cell(row=row, column=1).value = "Manifest version"
        ws.cell(row=row, column=2).value = manifest.get("manifest_version", 1)
        row += 1

        ws.cell(row=row, column=1).value = "Accepted proposal count"
        ws.cell(row=row, column=2).value = manifest.get("accepted_proposal_count", 0)
        row += 1

        ws.cell(row=row, column=1).value = "Manifest operation count"
        ws.cell(row=row, column=2).value = manifest.get("operation_count", 0)
        row += 2

        ws.cell(row=row, column=1).value = "Applied operations"
        row += 1
        ws.cell(row=row, column=1).value = "proposal_id"
        ws.cell(row=row, column=2).value = "op"
        ws.cell(row=row, column=3).value = "challenge_name"
        ws.cell(row=row, column=4).value = "previous_value"
        ws.cell(row=row, column=5).value = "new_value"
        row += 1

        for item in applied:
            ws.cell(row=row, column=1).value = item.get("proposal_id")
            ws.cell(row=row, column=2).value = item.get("op")
            ws.cell(row=row, column=3).value = item.get("challenge_name")
            ws.cell(row=row, column=4).value = item.get("previous_value")
            ws.cell(row=row, column=5).value = item.get("new_value")
            row += 1

        row += 1
        ws.cell(row=row, column=1).value = "Unresolved operations"
        row += 1
        ws.cell(row=row, column=1).value = "proposal_id"
        ws.cell(row=row, column=2).value = "op"
        ws.cell(row=row, column=3).value = "challenge_name"
        ws.cell(row=row, column=4).value = "reason"
        row += 1

        for item in unresolved:
            ws.cell(row=row, column=1).value = item.get("proposal_id")
            ws.cell(row=row, column=2).value = item.get("op")
            ws.cell(row=row, column=3).value = item.get("challenge_name")
            ws.cell(row=row, column=4).value = item.get("reason")
            row += 1

    def _headers(self, ws) -> dict[str, int]:
        headers: dict[str, int] = {}
        for idx, cell in enumerate(ws[1], start=1):
            value = cell.value
            if value is None:
                continue
            headers[str(value).strip()] = idx
        return headers