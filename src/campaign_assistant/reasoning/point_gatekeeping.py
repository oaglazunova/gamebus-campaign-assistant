from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class PointGatekeepingService:
    """
    First version of point/gatekeeping reasoning.

    It is deliberately conservative:
    - it reads actual campaign structure from the export
    - it uses explicit task role annotations when available
    - otherwise it falls back to simple heuristics
    - it produces explanations, warnings, and suggested fixes
    - it does not apply changes itself

    Current scope:
    - challenge-level target reachability
    - explicit/inferred gatekeeper analysis
    - basic maintenance-role awareness
    - structured warnings and suggestions

    Later scope:
    - at-risk return thresholds
    - stronger maintenance-vs-transition reasoning
    - comparison-aware point logic
    - patch generation
    """

    def analyze(
        self,
        *,
        campaign_file: str | Path,
        point_rules: dict[str, Any] | None = None,
        task_roles: list[dict[str, str]] | None = None,
    ) -> dict[str, Any]:
        point_rules = point_rules or {}
        task_roles = task_roles or []

        workbook = load_workbook(campaign_file, read_only=True, data_only=True)

        challenges = self._load_sheet_as_dicts(workbook, "challenges")
        tasks = self._load_sheet_as_dicts(workbook, "tasks")
        visualizations = self._load_sheet_as_dicts(workbook, "visualizations")

        workbook.close()

        challenge_by_id = {
            int(row["id"]): row for row in challenges if row.get("id") is not None
        }
        visualization_by_id = {
            str(row["id"]): row for row in visualizations if row.get("id") is not None
        }

        tasks_by_challenge: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for task in tasks:
            ch = task.get("challenge")
            if ch is None:
                continue
            try:
                tasks_by_challenge[int(ch)].append(task)
            except Exception:
                continue

        explicit_role_map = self._build_role_lookup(task_roles)

        findings: list[dict[str, Any]] = []
        warnings: list[str] = []
        suggestions: list[str] = []

        total_missing_targets = 0
        total_unreachable_targets = 0
        total_gatekeeper_warnings = 0
        total_maintenance_warnings = 0

        for challenge_id, challenge in challenge_by_id.items():
            challenge_name = str(challenge.get("name") or "")
            visualization_id = str(challenge.get("visualizations") or "")
            visualization_name = ""
            if visualization_id in visualization_by_id:
                visualization_name = str(
                    visualization_by_id[visualization_id].get("label")
                    or visualization_by_id[visualization_id].get("description")
                    or ""
                )

            challenge_tasks = tasks_by_challenge.get(challenge_id, [])
            if not challenge_tasks:
                continue

            target = challenge.get("target")
            target_value = self._as_number(target)

            task_summaries = []
            total_theoretical_points = 0.0

            explicit_gatekeepers = []
            explicit_maintenance = []
            inferred_gatekeepers = []

            for task in challenge_tasks:
                task_name = str(task.get("name") or "")
                points = self._as_number(task.get("points")) or 0.0
                max_times = self._as_number(task.get("max_times_fired")) or 0.0
                min_days_between = self._as_number(task.get("min_days_between_fire")) or 0.0

                theoretical_points = points * max_times
                total_theoretical_points += theoretical_points

                role = self._lookup_role(
                    explicit_role_map=explicit_role_map,
                    challenge_id=challenge_id,
                    challenge_name=challenge_name,
                    task_name=task_name,
                )

                if role == "gatekeeping":
                    explicit_gatekeepers.append(task_name)
                elif role == "maintenance":
                    explicit_maintenance.append(task_name)

                task_summaries.append(
                    {
                        "task_name": task_name,
                        "points": points,
                        "max_times_fired": max_times,
                        "min_days_between_fire": min_days_between,
                        "theoretical_points": theoretical_points,
                        "explicit_role": role,
                    }
                )

            # Simple heuristic gatekeeper inference:
            # the highest-point task(s), preferring low repetition opportunities.
            inferred_gatekeepers = self._infer_gatekeepers(task_summaries)

            challenge_warnings: list[str] = []
            challenge_suggestions: list[str] = []

            if target_value is None:
                total_missing_targets += 1
                challenge_warnings.append(
                    "Challenge has no target points defined."
                )
                challenge_suggestions.append(
                    "Define a target point threshold for this challenge."
                )
            else:
                if target_value > total_theoretical_points:
                    total_unreachable_targets += 1
                    challenge_warnings.append(
                        f"Target points ({target_value}) exceed the theoretical maximum reachable points ({total_theoretical_points})."
                    )
                    challenge_suggestions.append(
                        "Lower the target or increase achievable points/repetition allowances."
                    )

            uses_gatekeeping = point_rules.get("general", {}).get(
                "progression_should_require_gatekeeper", True
            )

            explicit_role_preferred = point_rules.get("gatekeeping", {}).get(
                "explicit_role_preferred", True
            )

            if uses_gatekeeping:
                if not explicit_gatekeepers:
                    total_gatekeeper_warnings += 1
                    if explicit_role_preferred:
                        challenge_warnings.append(
                            "No explicit gatekeeping task is marked for this challenge."
                        )
                        challenge_suggestions.append(
                            "Mark one or more gatekeeping tasks explicitly in task_roles.csv or in GameBus when supported."
                        )

                    if inferred_gatekeepers:
                        challenge_suggestions.append(
                            "Candidate gatekeeper task(s) inferred from the current configuration: "
                            + ", ".join(inferred_gatekeepers)
                        )

                # If target can be reached without the strongest inferred/explicit gatekeeper,
                # warn that progression may not actually depend on gatekeeping.
                gatekeeper_baseline = self._gatekeeper_baseline_points(
                    task_summaries=task_summaries,
                    explicit_gatekeepers=explicit_gatekeepers,
                    inferred_gatekeepers=inferred_gatekeepers,
                )

                if target_value is not None and gatekeeper_baseline is not None:
                    max_without_gatekeeper = total_theoretical_points - gatekeeper_baseline
                    if max_without_gatekeeper >= target_value:
                        total_gatekeeper_warnings += 1
                        challenge_warnings.append(
                            "Target appears reachable even without completing the effective gatekeeping task(s)."
                        )
                        challenge_suggestions.append(
                            "Increase target points, adjust gatekeeper points, or reduce non-gatekeeper contribution so progression truly depends on gatekeeping."
                        )

            maintenance_required = point_rules.get("maintenance", {}).get(
                "at_risk_return_should_depend_on_maintenance_tasks", False
            )

            challenge_name_lower = challenge_name.lower()
            looks_like_at_risk = "at risk" in challenge_name_lower

            if maintenance_required and looks_like_at_risk and not explicit_maintenance:
                total_maintenance_warnings += 1
                challenge_warnings.append(
                    "This looks like an at-risk challenge, but no explicit maintenance tasks are annotated."
                )
                challenge_suggestions.append(
                    "Mark maintenance tasks explicitly so return-from-risk logic can be checked more reliably."
                )

            if challenge_warnings:
                findings.append(
                    {
                        "challenge_id": challenge_id,
                        "challenge_name": challenge_name,
                        "visualization_name": visualization_name,
                        "target_points": target_value,
                        "theoretical_max_points": total_theoretical_points,
                        "explicit_gatekeepers": explicit_gatekeepers,
                        "explicit_maintenance": explicit_maintenance,
                        "inferred_gatekeepers": inferred_gatekeepers,
                        "warnings": challenge_warnings,
                        "suggestions": challenge_suggestions,
                    }
                )

        if total_missing_targets:
            warnings.append(
                f"{total_missing_targets} challenge(s) have no target points defined."
            )
        if total_unreachable_targets:
            warnings.append(
                f"{total_unreachable_targets} challenge(s) have target points that exceed the theoretical maximum."
            )
        if total_gatekeeper_warnings:
            warnings.append(
                f"{total_gatekeeper_warnings} challenge(s) may have gatekeeping problems or unclear gatekeeper definitions."
            )
        if total_maintenance_warnings:
            warnings.append(
                f"{total_maintenance_warnings} challenge(s) may have unclear maintenance-task annotations."
            )

        if findings:
            suggestions.append(
                "Review the highlighted challenges first; they are the most likely to require point/gatekeeping corrections."
            )
        if total_gatekeeper_warnings:
            suggestions.append(
                "Prefer explicit gatekeeping annotations over inference whenever possible."
            )
        if total_maintenance_warnings:
            suggestions.append(
                "Prefer explicit maintenance annotations for at-risk levels."
            )

        return {
            "summary": {
                "challenge_findings": len(findings),
                "missing_targets": total_missing_targets,
                "unreachable_targets": total_unreachable_targets,
                "gatekeeper_warnings": total_gatekeeper_warnings,
                "maintenance_warnings": total_maintenance_warnings,
            },
            "findings": findings,
            "warnings": warnings,
            "suggestions": suggestions,
        }

    def _load_sheet_as_dicts(self, workbook, sheet_name: str) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            return []

        ws = workbook[sheet_name]
        rows = ws.iter_rows(values_only=True)

        try:
            headers = next(rows)
        except StopIteration:
            return []

        headers = [str(h) if h is not None else "" for h in headers]
        result: list[dict[str, Any]] = []
        for row in rows:
            if row is None:
                continue
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else None
            result.append(item)
        return result

    def _build_role_lookup(self, task_roles: list[dict[str, str]]) -> list[dict[str, str]]:
        return task_roles

    def _lookup_role(
        self,
        *,
        explicit_role_map: list[dict[str, str]],
        challenge_id: int,
        challenge_name: str,
        task_name: str,
    ) -> str | None:
        challenge_id_str = str(challenge_id).strip().lower()
        challenge_name_norm = challenge_name.strip().lower()
        task_name_norm = task_name.strip().lower()

        for row in explicit_role_map:
            role = (row.get("role") or "").strip().lower()
            if not role:
                continue

            row_task_id = (row.get("task_id") or "").strip().lower()
            row_task_name = (row.get("task_name") or "").strip().lower()

            if row_task_id and row_task_id == challenge_id_str:
                return role

            if row_task_name and row_task_name == task_name_norm:
                return role

            if row_task_name and row_task_name == challenge_name_norm:
                return role

        return None

    def _infer_gatekeepers(self, task_summaries: list[dict[str, Any]]) -> list[str]:
        if not task_summaries:
            return []

        ranked = sorted(
            task_summaries,
            key=lambda x: (
                x.get("points", 0.0),
                -(x.get("max_times_fired", 0.0) or 0.0),
            ),
            reverse=True,
        )

        if not ranked:
            return []

        max_points = ranked[0].get("points", 0.0) or 0.0
        candidates = []
        for task in ranked:
            if (task.get("points", 0.0) or 0.0) == max_points:
                candidates.append(str(task["task_name"]))
        return candidates[:3]

    def _gatekeeper_baseline_points(
        self,
        *,
        task_summaries: list[dict[str, Any]],
        explicit_gatekeepers: list[str],
        inferred_gatekeepers: list[str],
    ) -> float | None:
        gatekeepers = set(explicit_gatekeepers or inferred_gatekeepers)
        if not gatekeepers:
            return None

        total = 0.0
        for task in task_summaries:
            if str(task["task_name"]) in gatekeepers:
                total += float(task.get("theoretical_points", 0.0) or 0.0)
        return total

    def _as_number(self, value: Any) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except Exception:
            return None