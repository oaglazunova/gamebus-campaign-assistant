from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from campaign_assistant.patches import PatchedExcelDraftGenerator


def _build_minimal_campaign(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    ws_challenges = wb.create_sheet("challenges")
    ws_challenges.append([
        "campaign",
        "id",
        "labels",
        "type",
        "name",
        "image",
        "description",
        "visualizations",
        "start",
        "end",
        "contenders",
        "is_initial_level",
        "target",
        "success_next",
        "evaluate_fail_every_x_minutes",
        "failure_next",
    ])
    ws_challenges.append([
        1,
        100,
        None,
        "TASKS_COLLECTION",
        "Challenge A",
        None,
        None,
        "200",
        None,
        None,
        None,
        1,
        20,
        101,
        10080,
        99,
    ])

    wb.save(path)


def test_excel_draft_generator_applies_set_target_points(tmp_path: Path):
    snapshot = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(snapshot)

    generator = PatchedExcelDraftGenerator(
        workspace_root=tmp_path / "workspace",
        request_id="req-001",
    )

    manifest = {
        "manifest_version": 1,
        "operation_count": 1,
        "accepted_proposal_count": 1,
        "operations": [
            {
                "op": "set_target_points",
                "proposal_id": "fix-1",
                "challenge_name": "Challenge A",
                "params": {
                    "target_points": 18,
                },
            }
        ],
        "skipped_proposals": [],
    }

    summary = generator.generate(
        snapshot_path=snapshot,
        manifest=manifest,
    )

    assert summary["applied_count"] == 1
    assert generator.draft_path.exists()

    wb = load_workbook(generator.draft_path)
    ws = wb["challenges"]
    assert ws.cell(row=2, column=13).value == 18  # target column
    assert PatchedExcelDraftGenerator.NOTES_SHEET_NAME in wb.sheetnames
    wb.close()


def test_excel_draft_generator_records_unresolved_operations(tmp_path: Path):
    snapshot = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(snapshot)

    generator = PatchedExcelDraftGenerator(
        workspace_root=tmp_path / "workspace",
        request_id="req-001",
    )

    manifest = {
        "manifest_version": 1,
        "operation_count": 1,
        "accepted_proposal_count": 1,
        "operations": [
            {
                "op": "annotate_gatekeeper",
                "proposal_id": "fix-2",
                "challenge_name": "Challenge A",
                "params": {
                    "candidate_gatekeepers": ["Task X"],
                },
            }
        ],
        "skipped_proposals": [],
    }

    summary = generator.generate(
        snapshot_path=snapshot,
        manifest=manifest,
    )

    assert summary["applied_count"] == 0
    assert summary["unresolved_count"] == 1

    wb = load_workbook(generator.draft_path)
    assert PatchedExcelDraftGenerator.NOTES_SHEET_NAME in wb.sheetnames
    notes_ws = wb[PatchedExcelDraftGenerator.NOTES_SHEET_NAME]
    assert notes_ws.cell(row=1, column=1).value == "GameBus Campaign Assistant Patch Notes"
    wb.close()


def test_excel_draft_generator_persists_summary_json(tmp_path: Path):
    snapshot = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(snapshot)

    generator = PatchedExcelDraftGenerator(
        workspace_root=tmp_path / "workspace",
        request_id="req-001",
    )

    manifest = {
        "manifest_version": 1,
        "operation_count": 0,
        "accepted_proposal_count": 0,
        "operations": [],
        "skipped_proposals": [],
    }

    summary = generator.generate(
        snapshot_path=snapshot,
        manifest=manifest,
    )

    assert generator.summary_path.exists()

    data = json.loads(generator.summary_path.read_text(encoding="utf-8"))
    assert data["request_id"] == "req-001"
    assert data["draft_path"] == str(generator.draft_path)