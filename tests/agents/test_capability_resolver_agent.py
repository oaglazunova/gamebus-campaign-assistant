from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook

from campaign_assistant.agents.capability_resolver import CapabilityResolverAgent
from campaign_assistant.orchestration.models import AgentContext


def _build_minimal_campaign(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "challenges"
    ws.append(
        [
            "campaign",
            "id",
            "labels",
            "type",
            "name",
            "image",
            "description",
            "visualizations",
            "start",
            "end",
            "contenders",
            "is_initial_level",
            "target",
            "success_next",
            "evaluate_fail_every_x_minutes",
            "failure_next",
        ]
    )
    ws.append(
        [
            1,
            100,
            None,
            "TASKS_COLLECTION",
            "Challenge A",
            None,
            None,
            "200",
            None,
            None,
            None,
            1,
            20,
            101,
            10080,
            99,
        ]
    )

    ws_tasks = wb.create_sheet("tasks")
    ws_tasks.append(["id", "challenge", "name", "points", "max_times_fired"])
    ws_tasks.append([1, 100, "Walk 20 minutes", 10, 2])

    ws_waves = wb.create_sheet("waves")
    ws_waves.append(["id", "start", "end"])

    ws_groups = wb.create_sheet("groups")
    ws_groups.append(["id", "name"])

    wb.save(path)


def _make_context(tmp_path: Path) -> AgentContext:
    campaign_file = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(campaign_file)

    workspace_root = tmp_path / "workspace"
    metadata_dir = workspace_root / "metadata"
    metadata_dir.mkdir(parents=True, exist_ok=True)

    with (metadata_dir / "campaign_profile.json").open("w", encoding="utf-8") as fh:
        json.dump(
            {
                "capabilities": {
                    "uses_progression": True,
                    "uses_gatekeeping": True,
                    "uses_ttm": False,
                }
            },
            fh,
            indent=2,
        )

    with (metadata_dir / "task_roles.csv").open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["task_id", "task_name", "role", "notes"])
        writer.writeheader()
        writer.writerow(
            {
                "task_id": "1",
                "task_name": "Walk 20 minutes",
                "role": "gatekeeping",
                "notes": "Example",
            }
        )

    return AgentContext(
        request_id="req-cap-001",
        file_path=campaign_file,
        selected_checks=["ttm"],
        export_excel=False,
        workspace_id="ws-cap",
        workspace_root=workspace_root,
        snapshot_id="snap-001",
        analysis_profile={},
        point_rules={},
        task_roles=[],
        evidence_index={},
        shared={},
    )


def test_capability_resolver_loads_metadata_bundle_and_summary(tmp_path: Path):
    ctx = _make_context(tmp_path)

    agent = CapabilityResolverAgent()
    response = agent.run(ctx)

    assert response.success is True
    assert "metadata_bundle" in ctx.shared
    assert "capability_summary" in ctx.shared

    summary = ctx.shared["capability_summary"]
    assert summary["capabilities"]["uses_progression"] is True
    assert summary["capabilities"]["uses_gatekeeping"] is True
    assert summary["capabilities"]["uses_ttm"] is False
    assert summary["task_role_count"] == 1


def test_capability_resolver_sets_active_modules(tmp_path: Path):
    ctx = _make_context(tmp_path)

    agent = CapabilityResolverAgent()
    agent.run(ctx)

    summary = ctx.shared["capability_summary"]
    modules = summary["active_modules"]

    assert modules["structural_checks"] is True
    assert modules["point_gatekeeping_checks"] is True
    assert modules["ttm_checks"] is False
    assert modules["content_fix_suggestions"] is True