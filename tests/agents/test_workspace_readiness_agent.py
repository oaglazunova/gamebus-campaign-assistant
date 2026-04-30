from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from campaign_assistant.agents.privacy_guardian import PrivacyGuardianAgent
from campaign_assistant.agents.workspace_readiness import WorkspaceReadinessAgent
from campaign_assistant.orchestration.models import AgentContext


def _build_minimal_campaign(path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "challenges"
    ws.append(
        [
            "campaign",
            "id",
            "labels",
            "type",
            "name",
            "image",
            "description",
            "visualizations",
            "start",
            "end",
            "contenders",
            "is_initial_level",
            "target",
            "success_next",
            "evaluate_fail_every_x_minutes",
            "failure_next",
        ]
    )
    ws.append(
        [
            1,
            100,
            None,
            "TASKS_COLLECTION",
            "Challenge A",
            None,
            None,
            "200",
            None,
            None,
            None,
            1,
            20,
            100,
            10080,
            100,
        ]
    )

    ws_tasks = wb.create_sheet("tasks")
    ws_tasks.append(
        [
            "id",
            "challenge",
            "name",
            "points",
            "max_times_fired",
            "min_days_between_fire",
        ]
    )
    ws_tasks.append([1, 100, "Walk 20 minutes", 10, 2, 1])

    ws_waves = wb.create_sheet("waves")
    ws_waves.append(["id", "start", "end"])

    ws_groups = wb.create_sheet("groups")
    ws_groups.append(["id", "name"])

    wb.save(path)


def _make_context(tmp_path: Path) -> AgentContext:
    campaign_file = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(campaign_file)

    workspace_root = tmp_path / "workspace"
    metadata_dir = workspace_root / "metadata"
    metadata_dir.mkdir(parents=True, exist_ok=True)
    (metadata_dir / "campaign_profile.json").write_text("{}", encoding="utf-8")

    ctx = AgentContext(
        request_id="req-readiness-001",
        file_path=campaign_file,
        selected_checks=[],
        export_excel=False,
        workspace_id="ws-readiness",
        workspace_root=workspace_root,
        snapshot_id="snap-readiness-001",
        analysis_profile={},
        point_rules={},
        task_roles=[],
        evidence_index={},
        shared={},
    )
    ctx.shared["capability_summary"] = {
        "capabilities": {"uses_progression": True},
        "task_role_count": 0,
    }
    return ctx


def test_workspace_readiness_agent_stores_readiness_and_merges_into_capability_summary(tmp_path: Path):
    ctx = _make_context(tmp_path)

    PrivacyGuardianAgent().run(ctx)
    response = WorkspaceReadinessAgent().run(ctx)

    assert response.success is True
    assert "workspace_readiness" in ctx.shared
    assert "workspace_readiness" in ctx.shared["capability_summary"]