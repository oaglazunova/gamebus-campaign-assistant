from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from campaign_assistant.reasoning import PointGatekeepingService


def _build_minimal_campaign(path: Path, target_value=20) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_challenges = wb.create_sheet("challenges")
    ws_challenges.append([
        "campaign",
        "id",
        "labels",
        "type",
        "name",
        "image",
        "description",
        "visualizations",
        "start",
        "end",
        "contenders",
        "is_initial_level",
        "target",
        "success_next",
        "evaluate_fail_every_x_minutes",
        "failure_next",
    ])
    ws_challenges.append([
        1,
        100,
        None,
        "TASKS_COLLECTION",
        "[Skilled] Test Challenge",
        None,
        None,
        "200",
        None,
        None,
        None,
        1,
        target_value,
        101,
        10080,
        99,
    ])

    ws_tasks = wb.create_sheet("tasks")
    ws_tasks.append([
        "challenge",
        "name",
        "description",
        "image",
        "video",
        "h5p_slug",
        "story_url",
        "max_times_fired",
        "min_days_between_fire",
        "activityscheme_default",
        "activityschemes_allowed",
        "image_required",
        "conditions",
        "points",
        "dataproviders",
    ])
    ws_tasks.append([
        100,
        "Gatekeeper task",
        None,
        None,
        None,
        None,
        None,
        1,
        7,
        "WATCH_STORY",
        "WATCH_STORY",
        0,
        None,
        10,
        "GameBus Studio",
    ])
    ws_tasks.append([
        100,
        "Regular task",
        None,
        None,
        None,
        None,
        None,
        5,
        1,
        "WATCH_STORY",
        "WATCH_STORY",
        0,
        None,
        2,
        "GameBus Studio",
    ])

    ws_visualizations = wb.create_sheet("visualizations")
    ws_visualizations.append([
        "campaign",
        "id",
        "description",
        "label",
        "icon",
        "href",
        "target_is_blank",
        "assignment",
        "for_organizer",
        "in_menu",
        "menu_order",
        "in_tabbar",
        "tabbar_order",
        "settings",
        "visualization",
        "wave",
        "groups",
    ])
    ws_visualizations.append([
        1,
        200,
        "Test Visualization",
        "Test Visualization",
        None,
        None,
        1,
        "INDIVIDUAL",
        0,
        1,
        1,
        1,
        1,
        None,
        "level",
        300,
        "400",
    ])

    wb.save(path)


def test_point_gatekeeping_uses_explicit_gatekeeper_role(tmp_path: Path):
    campaign = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(campaign, target_value=20)

    service = PointGatekeepingService()
    result = service.analyze(
        campaign_file=campaign,
        point_rules={
            "general": {"progression_should_require_gatekeeper": True},
            "gatekeeping": {"explicit_role_preferred": True},
            "maintenance": {"at_risk_return_should_depend_on_maintenance_tasks": False},
        },
        task_roles=[
            {
                "task_id": "",
                "task_name": "Gatekeeper task",
                "role": "gatekeeping",
                "notes": "",
            }
        ],
    )

    # The challenge should be analyzed; no requirement that it must be broken.
    assert "summary" in result
    assert "findings" in result
    # In this synthetic example, the gatekeeper is explicit and target=20, while
    # max without gatekeeper is only 10, so it should not warn about bypassing gatekeeping.
    joined_warnings = " ".join(result["warnings"])
    assert "gatekeeping problems" not in joined_warnings.lower()


def test_point_gatekeeping_warns_when_target_missing(tmp_path: Path):
    campaign = tmp_path / "campaign_missing_target.xlsx"
    _build_minimal_campaign(campaign, target_value=None)

    service = PointGatekeepingService()
    result = service.analyze(
        campaign_file=campaign,
        point_rules={
            "general": {"progression_should_require_gatekeeper": True},
            "gatekeeping": {"explicit_role_preferred": True},
            "maintenance": {"at_risk_return_should_depend_on_maintenance_tasks": False},
        },
        task_roles=[],
    )

    assert result["summary"]["missing_targets"] == 1
    assert any("no target points defined" in w.lower() for w in result["warnings"])


def test_point_gatekeeping_warns_when_target_exceeds_theoretical_max(tmp_path: Path):
    campaign = tmp_path / "campaign_unreachable.xlsx"
    _build_minimal_campaign(campaign, target_value=1000)

    service = PointGatekeepingService()
    result = service.analyze(
        campaign_file=campaign,
        point_rules={
            "general": {"progression_should_require_gatekeeper": True},
            "gatekeeping": {"explicit_role_preferred": True},
            "maintenance": {"at_risk_return_should_depend_on_maintenance_tasks": False},
        },
        task_roles=[],
    )

    assert result["summary"]["unreachable_targets"] == 1
    assert any("exceed the theoretical maximum" in w.lower() for w in result["warnings"])