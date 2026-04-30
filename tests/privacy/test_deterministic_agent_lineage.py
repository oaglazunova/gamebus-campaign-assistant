from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook

from campaign_assistant.agents.capability_resolver import CapabilityResolverAgent
from campaign_assistant.agents.privacy_guardian import PrivacyGuardianAgent
from campaign_assistant.agents.structural_change import StructuralChangeAgent
from campaign_assistant.orchestration.models import AgentContext


def _build_minimal_campaign(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "challenges"
    ws.append(
        [
            "campaign",
            "id",
            "labels",
            "type",
            "name",
            "image",
            "description",
            "visualizations",
            "start",
            "end",
            "contenders",
            "is_initial_level",
            "target",
            "success_next",
            "evaluate_fail_every_x_minutes",
            "failure_next",
        ]
    )
    ws.append(
        [
            1,
            100,
            None,
            "TASKS_COLLECTION",
            "Challenge A",
            None,
            None,
            "200",
            None,
            None,
            None,
            1,
            20,
            101,
            10080,
            99,
        ]
    )

    ws_tasks = wb.create_sheet("tasks")
    ws_tasks.append(["id", "challenge", "name", "points", "max_times_fired"])
    ws_tasks.append([1, 100, "Walk 20 minutes", 10, 2])

    ws_waves = wb.create_sheet("waves")
    ws_waves.append(["id", "start", "end"])

    ws_groups = wb.create_sheet("groups")
    ws_groups.append(["id", "name"])

    wb.save(path)


def _make_context(tmp_path: Path, *, selected_checks: list[str] | None = None) -> AgentContext:
    campaign_file = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(campaign_file)

    workspace_root = tmp_path / "workspace"
    metadata_dir = workspace_root / "metadata"
    metadata_dir.mkdir(parents=True, exist_ok=True)

    with (metadata_dir / "campaign_profile.json").open("w", encoding="utf-8") as fh:
        json.dump(
            {
                "capabilities": {
                    "uses_progression": True,
                    "uses_gatekeeping": True,
                    "uses_ttm": False,
                }
            },
            fh,
            indent=2,
        )

    with (metadata_dir / "task_roles.csv").open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["task_id", "task_name", "role", "notes"])
        writer.writeheader()
        writer.writerow(
            {
                "task_id": "1",
                "task_name": "Walk 20 minutes",
                "role": "gatekeeping",
                "notes": "Example",
            }
        )

    return AgentContext(
        request_id="req-det-lineage-001",
        file_path=campaign_file,
        selected_checks=selected_checks or [],
        export_excel=False,
        workspace_id="ws-det-lineage",
        workspace_root=workspace_root,
        snapshot_id="snap-det-lineage-001",
        analysis_profile={},
        point_rules={},
        task_roles=[],
        evidence_index={},
        shared={},
    )


def test_capability_resolver_records_deterministic_lineage(tmp_path: Path):
    ctx = _make_context(tmp_path)

    PrivacyGuardianAgent().run(ctx)
    response = CapabilityResolverAgent().run(ctx)

    assert response.success is True

    audit_log = ctx.shared["privacy_state"]["audit_log"]
    resolver_events = [e for e in audit_log if e.get("agent_name") == "capability_resolver_agent"]

    assert any(e["event"] == "agent_run_started" for e in resolver_events)
    assert any(e["event"] == "asset_access_recorded" for e in resolver_events)
    assert any(e["event"] == "agent_completed" for e in resolver_events)

    access_events = [e for e in resolver_events if e["event"] == "asset_access_recorded"]
    assert any(e["asset_id"] == "campaign_workbook" for e in access_events)


def test_structural_change_records_deterministic_lineage(tmp_path: Path):
    ctx = _make_context(tmp_path, selected_checks=[])

    PrivacyGuardianAgent().run(ctx)
    CapabilityResolverAgent().run(ctx)
    response = StructuralChangeAgent().run(ctx)

    assert response.success is True

    audit_log = ctx.shared["privacy_state"]["audit_log"]
    structural_events = [e for e in audit_log if e.get("agent_name") == "structural_change_agent"]

    assert any(e["event"] == "agent_run_started" for e in structural_events)
    assert any(e["event"] == "asset_access_recorded" for e in structural_events)
    assert any(e["event"] == "agent_completed" for e in structural_events)

    access_events = [e for e in structural_events if e["event"] == "asset_access_recorded"]
    assert any(e["asset_id"] == "campaign_workbook" for e in access_events)