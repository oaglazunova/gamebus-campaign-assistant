from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook

from campaign_assistant.metadata.adapters.merged import load_merged_metadata_bundle


def _build_minimal_campaign(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    wb.title = "challenges"
    ws.append(
        [
            "campaign",
            "id",
            "labels",
            "type",
            "name",
            "image",
            "description",
            "visualizations",
            "start",
            "end",
            "contenders",
            "is_initial_level",
            "target",
            "success_next",
            "evaluate_fail_every_x_minutes",
            "failure_next",
        ]
    )
    ws.append(
        [
            1,
            100,
            None,
            "TASKS_COLLECTION",
            "Challenge A",
            None,
            None,
            "200",
            None,
            None,
            None,
            1,
            20,
            101,
            10080,
            99,
        ]
    )

    ws_tasks = wb.create_sheet("tasks")
    ws_tasks.append(["id", "challenge", "name", "points", "max_times_fired"])
    ws_tasks.append([1, 100, "Walk 20 minutes", 10, 2])

    ws_waves = wb.create_sheet("waves")
    ws_waves.append(["id", "start", "end"])

    ws_groups = wb.create_sheet("groups")
    ws_groups.append(["id", "name"])

    wb.save(path)


def test_merged_metadata_prefers_sidecar_over_inferred(tmp_path: Path):
    campaign_path = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(campaign_path)

    workspace_root = tmp_path / "workspace"
    metadata_dir = workspace_root / "metadata"
    metadata_dir.mkdir(parents=True, exist_ok=True)

    with (metadata_dir / "campaign_profile.json").open("w", encoding="utf-8") as fh:
        json.dump(
            {
                "capabilities": {
                    "uses_progression": False,
                    "uses_ttm": True,
                    "uses_gatekeeping": True,
                }
            },
            fh,
        )

    bundle = load_merged_metadata_bundle(
        file_path=campaign_path,
        workspace_root=workspace_root,
    )

    # inferred progression would likely be True because success/failure transitions exist,
    # but sidecar should override it
    assert bundle.capabilities.uses_progression is False
    assert bundle.capabilities.uses_ttm is True
    assert bundle.capabilities.uses_gatekeeping is True


def test_merged_metadata_loads_task_roles(tmp_path: Path):
    campaign_path = tmp_path / "campaign.xlsx"
    _build_minimal_campaign(campaign_path)

    workspace_root = tmp_path / "workspace"
    metadata_dir = workspace_root / "metadata"
    metadata_dir.mkdir(parents=True, exist_ok=True)

    with (metadata_dir / "task_roles.csv").open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["task_id", "task_name", "role", "notes"])
        writer.writeheader()
        writer.writerow(
            {
                "task_id": "",
                "task_name": "Walk 20 minutes",
                "role": "gatekeeping",
                "notes": "Manual annotation",
            }
        )

    bundle = load_merged_metadata_bundle(
        file_path=campaign_path,
        workspace_root=workspace_root,
    )

    assert len(bundle.task_roles) == 1
    assert bundle.task_roles[0].role == "gatekeeping"